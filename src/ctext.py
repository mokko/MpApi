"""

Import C-Texts to RIA. This will be a two-step process.
(1) We parse A. Schäfer's Excel sheets (one at a time) and write the required
    info into a new Excel file sanitizing it somewhat on the way; 
    in this step we also look up the objIds from RIA (plus possibly other 
    information).
(2) we manually proof-reading the new Excel and potentially correct it    
(3) we upload stuff from the new Excel to RIA

Let's find good names 
(A) for the products: 
- A.Schäfer's files are the __original_Excels__
- The Excel sheets that I create are __the_proofing_files__

(B) for the stages:
- initial conversion
- manual proofing
- upload to RIA

USAGE
    # step 1
    ctext --input original.xslx --output m39-proof.xslx  
    # step 2 happens in Excel (without this program)
    # step 3
    ctext --upload m39-proof.xslx  
"""

import argparse
from mpapi.search import Search
from mpapi.client import MpApi
from mpapi.module import Module
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Alignment, Font
import pprint
from typing import Optional

with open("credentials.py") as f:
    exec(f.read())


class Ctext:
    def __init__(self, *, user: str, pw: str, baseURL: str) -> None:
        self.client = MpApi(baseURL=baseURL, user=user, pw=pw)

    def _ingestSheet(self, sheet) -> None:
        """
        Process the original Excel (in sheet) and prepare the proofing Excel;
        as a side-effect stores proofing excel to self.wb2 (and self.ws2).

        EXPECTS
        - sheet: openpyxl sheet

        RETURNS
        - nothing
        """
        print(f"*max row: {self.ws2.max_row}")  # 1-based
        ws2 = self.ws2
        new_row = ws2.max_row
        rno = 11  # 1-based
        for old_row in sheet.iter_rows(min_row=12, max_col=30):
            rno += 1
            new_row += 1
            identNrXLS = str(sheet[f"E{rno}"].value).strip()
            if identNrXLS == "./.":
                break  # are neglecting information in the rest of the line?

            print(f"\n***{new_row} identNr {identNrXLS}")

            # mapping original Excel fields to internal fields
            zeichen = sheet[f"W{rno}"].value
            künstler_de = sheet[f"o{rno}"].value
            künstler_en = sheet[f"p{rno}"].value
            titel_de = sheet[f"q{rno}"].value
            titel_en = sheet[f"r{rno}"].value
            kennung_de = sheet[f"U{rno}"].value
            kennung_en = sheet[f"V{rno}"].value

            # character translations: strange bullets
            mt = str.maketrans("#", "\u00B7")  # U+00B7 2022
            kennung_de = kennung_de.translate(mt)
            kennung_en = kennung_en.translate(mt)

            # DEBUGGING
            # print(f"z {zeichen}")
            # print(f"kü {künstler_de} || {künstler_en}")
            # print(f"ti {titel_de} || {titel_en}")
            # print(f"ke {kennung_de}")
            # print(f"ke {kennung_en}")

            # deal with None and join fields
            de = ""
            if zeichen is not None:
                de += zeichen
            if künstler_de is not None and künstler_de != "./.":
                de += f"\n{künstler_de}"
            if titel_de is not None:
                de += f"\n\n{titel_de}"
            if kennung_de is not None:
                de += f"\n\n{kennung_de}"

            en = ""
            if zeichen is not None:
                en += zeichen
            if künstler_en is not None and künstler_en != "./.":
                en += f"\n{künstler_en}"
            if titel_en is not None:
                en += f"\n\n{titel_en}"
            if kennung_en is not None:
                en += f"\n\n{kennung_en}"

            # todo: fix this html, when we can see it in RIA
            de_html = "<html>"
            if zeichen is not None:
                zeichen = zeichen.strip()
                de_html += zeichen + "<br/>"
            if künstler_de == "./.":
                künstler_de = None
            if künstler_de is not None:
                künstler_de = künstler_de.strip()
                de_html += künstler_de + "<br/>"
            if titel_de is not None:
                titel_de = titel_de.strip()
                de_html += f"<p><b>{titel_de}</b></p>"
            if kennung_de is not None:
                kennung_de = kennung_de.strip()
                de_html += f"<p>{kennung_de}</p>"
            de_html += "</html>"

            en_html = "<html>"
            if zeichen is not None:
                en_html += zeichen + "<br/>"
            if künstler_en == "./.":
                künstler_en = None
            if künstler_en is not None:
                künstler_en = künstler_en.strip()
                en_html += künstler_en + "<br/>"
            if titel_en is not None:
                titel_en = titel_en.strip()
                en_html += f"<p><b>{titel_en}</b></p>"
            if kennung_de is not None:
                kennung_en = kennung_en.strip()
                en_html += f"<p>{kennung_en}</p>"
            en_html += "</html>"

            # mapping internal fields to new Excel fields
            ws2[f"C{new_row}"] = identNrXLS
            ws2[f"D{new_row}"] = de
            ws2[f"D{new_row}"].style = "wrap"
            ws2[f"E{new_row}"] = en
            ws2[f"E{new_row}"].style = "wrap"
            ws2[f"F{new_row}"] = sheet[f"B{rno}"].value  # element
            ws2[f"G{new_row}"] = sheet.title  # Blatt
            ws2[f"H{new_row}"] = self.origFn  # Datei
            ws2[f"I{new_row}"] = de_html
            ws2[f"I{new_row}"].style = "wrap"
            ws2[f"J{new_row}"] = en_html
            ws2[f"J{new_row}"].style = "wrap"
            ws2[f"K{new_row}"] = zeichen
            ws2[f"L{new_row}"] = künstler_de
            ws2[f"M{new_row}"] = künstler_en
            ws2[f"N{new_row}"] = titel_de
            ws2[f"O{new_row}"] = titel_en
            ws2[f"P{new_row}"] = kennung_de
            ws2[f"P{new_row}"].style = "wrap"
            ws2[f"Q{new_row}"] = kennung_en
            ws2[f"Q{new_row}"].style = "wrap"

    def _initProof(self) -> None:
        """
        Initialize rewrite Excel (self.ws2). As a side effect: creates a new
        Excel in memory (self.wb2 with active sheet in self.ws2)

        Expects nothing and returns nothing
        """
        self.wb2 = Workbook()  # new Excel
        ws2 = self.wb2.active
        self.ws2 = ws2
        ws2["A1"] = "objId(RIA)"
        ws2["B1"] = "Bereich(RIA)"
        ws2["C1"] = "IdentNr(xlsx)"
        ws2["D1"] = "C-Text de"
        ws2.column_dimensions["D"].width = 45
        ws2["E1"] = "C-Text en"
        ws2.column_dimensions["E"].width = 45
        ws2["F1"] = "Element"
        ws2["G1"] = "Blatt"
        ws2["H1"] = "Datei"
        ws2["I1"] = "de_html"
        ws2.column_dimensions["I"].width = 45
        ws2["J1"] = "en_html"
        ws2.column_dimensions["J"].width = 45
        ws2["K1"] = "zeichen"
        ws2["L1"] = "künstler_de"
        ws2["M1"] = "künstler_en"
        ws2["N1"] = "titel_de"
        ws2["O1"] = "titel_en"
        ws2["P1"] = "kennung_de"
        ws2.column_dimensions["P"].width = 45
        ws2["Q1"] = "kennung_en"
        ws2.column_dimensions["Q"].width = 45

        ns = NamedStyle(name="wrap")
        ns.alignment = Alignment(vertical="top", wrap_text=True)
        self.wb2.add_named_style(ns)

    def _lookup(self, *, identNr: str, limit: int = -1) -> Optional[str]:
        """
        Some unicode/latin-1 mess is happening here...

        UnicodeEncodeError: 'latin-1' codec can't encode character '\u2014' in
        position 453: Body ('—') is not valid Latin-1. Use body.encode('utf-8')
        if you want to send it encoded in UTF-8.
        """

        # print(f"-----------LOOKING UP objId for identNr {identNr}")
        q = Search(module="Object", limit=limit)
        q.AND()
        q.addCriterion(
            operator="equalsField",
            field="ObjObjectNumberVrt",  # ObjObjectNumberVrt
            value=identNr,
        )
        q.OR()
        AKuList = [
            "AKuArchivOAK",
            "AKuArchivSSOZ",
            #    "AKuFotoarchiv",
            "AKuKriegsverlusteOAK",
            "AKuKriegsverlusteSSOZ",
            "AKuSudSudostundZentralasien",
            "AKuOstasiatischeKunst",
        ]
        for orgUnit in AKuList:
            q.addCriterion(
                operator="equalsField",  # notEqualsTerm
                field="__orgUnit",  # __orgUnit is not allowed in Zetcom's own search.xsd
                value=orgUnit,
            )
        q.addField(field="__id")
        q.addField(field="__orgUnit")
        q.validate(mode="search")
        # print(q.toString())
        m = self.client.search2(query=q)
        # print (m.toString())
        size = m.actualSize(module="Object")
        if size == 1:
            try:
                newID = m.xpath(
                    "/m:application/m:modules/m:module[@name = 'Object']/m:moduleItem/@id"
                )[0]
            except:
                print("WARN: xpath fails; zero results?")
                newID = None
            try:
                bereich = m.xpath(
                    """/m:application/m:modules/m:module[
                        @name = 'Object'
                    ]/m:moduleItem/m:systemField[
                        @name = '__orgUnit'
                    ]/m:value/text()
                    """
                )[0]
            except:
                bereich = None
            if newID is not None and bereich is None:
                raise ValueError("ERROR: bereich should exist")
            # print (f"***bereich:{bereich}")
        else:
            print(f"WARN: multiple or zero results {size}")
            newID = None
            bereich = None
        return newID, bereich

    def _upload(self, *, row: dict) -> None:
        """
        Uploads contents of row to RIA.

        expects a row as dictionary
        returns nothing much

        First we write only C-Text in Texte, later we might also fill up
        onlineBeschreibung with parts of the info here.
        """
        self._upTexte(row=row)
        self._upTexteOnline(row=row)

        pp = pprint.PrettyPrinter(indent=4)
        pp.pprint(row)

    def _upTexte(self, *, row: dict):
        """
        DF: a label is a record in field cluster Texte with the qualifier
            typ = "Label"

        Mapping Vorschlag
            Typ: Label
            Ausstellung: keine
            Anlass: Ausstellung
            Sprache: de
            Text HTML: <gesamter formatierter Label-Text>

        Update Policy:
        - dont overwrite existing labels; i.e. if label exists already, do
          nothing so as to make recurring executions harmless
        - If label doesn't exist yet (not a single label exists yet), add new
          label according to mapping above.

        1. search: We have to get the current record from RIA; do we only get records that dont
        have label yet?
        x. loop through results -> not applicable here since we only get a single result. Dont
        need to check that since objID are reasonbly unique.
        2. Since the search is strict and weeds out records with existing
        label, we only have one change: that we create a completely new label

        """
        print("_upTexte")

    def _upTexteOnline(self, *, row: dict):
        """
        DF: onlineText is an record/entry in Texte Online with typ = "Online Beschreibung"

        Typically, if an onlineText exists already, we leave it alone. The only exception is
        if the text contains only of "SM8HF". In this case, we add our new text anyways.

        - search record for objId, get texte cluster
        - investigate onlineTexte cluster; if doesn't exist or only contains [SM8HF], continue
        - if cluster doesn't exist yet, create new texteOnline
        - if cluster exists and only includes [SM8HF], add new text
        """
        pass

    def ingest(self, *, origFn: str, proofFn: str) -> None:
        """
        Rewrite original Excel input to the proofing Excel and save at provided
        path.

        Expects
        - origFn: path to original Excel file
        - proofFn: path for rewritten proofing Excel file

        Returns nothing.
        """
        wb1 = load_workbook(origFn)  # original
        self.origFn = origFn  # used in _ingestSheet
        self._initProof()

        # writes multiple sheets from origin into one sheet at proofFn
        for sheet in wb1:
            if sheet.title != "Info zur Datei":
                self._ingestSheet(sheet)
        openpyxl.writer.excel.save_workbook(self.wb2, proofFn)

    def riaLookup(self, *, proofFn: str) -> None:
        """
        For every identNr from Excel, look up objId and IdentNr from RIA and
        write the results back to the proofing Excel (self.ws2). Save proofing
        Excel after every lookup, so that in case of interruptions, requests
        don't have to be repeated.

        Expects
        - proofFn: loation for proofing Excel

        Returns nothing
        """
        if not hasattr(self, "wb2"):
            self.wb2 = load_workbook(proofFn)
            self.ws2 = self.wb2.active

        ws2 = self.ws2
        max_row = self.ws2.max_row
        print("*Entering ria lookup")
        print(f"**max row: {max_row}")  # 1-based

        for rno in range(2, max_row + 1):
            riaObjId = ws2[f"A{rno}"].value
            xslIdentNr = ws2[f"C{rno}"].value
            if riaObjId is None:
                newID, bereich = self._lookup(identNr=xslIdentNr, limit=1)
                if newID is not None:
                    ws2[f"A{rno}"] = newID
                if bereich is not None:
                    ws2[f"B{rno}"] = bereich
                    openpyxl.writer.excel.save_workbook(self.wb2, proofFn)
                    print(f"***row #{rno} {riaObjId} {xslIdentNr} -> {newID} {bereich}")

    def upload(self, *, proofFn: str) -> None:
        """
        Let's keep the Excel stuff here and move the mpapi stuff into a extra
        method.
        """
        self.wb2 = load_workbook(proofFn)
        ws2 = self.wb2.active
        max_row = ws2.max_row
        print(f"*Entering UPLOAD step max row: {max_row}")  # 1-based

        # row to dict
        rno = 1  # one-based row counter
        headings = []  # zero-based
        for row in ws2.iter_rows(min_row=1, max_col=17):  # leave out params here?
            riaObjId = row[0].value
            print(f"riaObjId {riaObjId}")
            cno = 0  # zero-based cell counter
            rowDict = {}
            for cell in row:
                if rno == 1:
                    headings.append(cell.value)
                else:
                    rowDict[headings[cno]] = cell.value
                    cno += 1
            # only attempt RIA update if an objId has been identified
            if riaObjId is not None:
                self._upload(row=rowDict)
            rno += 1


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Upload c-texts in two steps")
    parser.add_argument("-o", "--orig", help="path to Excel input file")
    parser.add_argument("-p", "--proof", help="path to proofing Excel)")
    parser.add_argument("-l", "--lookup", help="path to proofing Excel")
    parser.add_argument("-u", "--upload", help="path to proofing Excel")
    args = parser.parse_args()

    ct = Ctext(user=user, pw=pw, baseURL=baseURL)
    if args.orig is not None:
        ct.ingest(origFn=args.orig, proofFn=args.proof)
        ct.riaLookup(proofFn=args.proof)

    if args.lookup:
        ct.riaLookup(proofFn=args.lookup)

    if args.upload is not None:
        ct.upload(proofFn=args.upload)
