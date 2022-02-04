"""
du - download/upload allows downloading data from MuseumPlus to an excel sheet
and uploading it back after changes have been made to the excel 

THIS IS THE FIRST PROOF OF CONCEPT VERSION. WORKS ONLY WITH dataFields

NOT FULLY TESTED.

WORKFLOW
1. You prepare a Excel sheet, e.g. using RIA.
2. You use du to fill the fields of the Excel with content from RIA
3. You make manual changes to the Excel sheet
4. You upload the changes to the RIA using du

FORMAT OF THE EXCEL FILE
   A        B                      C          ...
1  Download 20220130T17:50:00:000
2  Object   Fieldname1             Fieldname2 ...
3  123456
4  123347

The excel format in words: 
* Exceptions at the beginning  
    B1 provides the date of the last download; du enters the date automatically 
    during download
    A2 module type to which ids belong, e.g. Object
* Rest of column A lists the IDs
* You may also list respective IdentNr in the second row for your orientation.
* The rest of the second row should be the fieldnames you want to download. 
  You should use the internal fieldnames (probably using a special notation 
  which is not determined yet.)
* at this point du works only the first Excel sheet, but this may change in the
  future.     
* sheet names are ignored at this point

COMMAND LINE INTERFACE (subject to change)
    du -c down -i path/to/example.xlsx  # changes Excel so backup frequently
    du -c up -i path/to/example.xlsx

UPDATES to RIA are only written
- if value from Excel is different from online value
- if online record has not been changed since the download

KNOWN LIMITATIONS
* only accepts *.xlsx format, not *.xls -> Reformat using Excel
* If Excel file is still opened in Excel, it is locked and this script can't 
  run (Permission denied). -> Close Excel 
* there is a Unix command with the same name
* currently we only allow the modules Object, Person, Multimedia
"""

from datetime import datetime
import dateutil.parser
from lxml import etree  # type: ignore
from mpapi.client import MpApi
from mpapi.sar import Sar
from mpapi.module import Module
from mpapi.search import Search

import openpyxl  # type: ignore
from openpyxl import load_workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from openpyxl.styles.colors import Color  # type: ignore

NSMAP = {
    "s": "http://www.zetcom.com/ria/ws/module/search",
    "m": "http://www.zetcom.com/ria/ws/module",
}

known_modules = ["Object", "Person", "Multimedia"]


class Du:
    def __init__(
        self, *, cmd: str, Input: str, baseURL: str, user: str, pw: str
    ) -> None:
        self.api = MpApi(baseURL=baseURL, user=user, pw=pw)
        self.sar = Sar(baseURL=baseURL, user=user, pw=pw)
        self.Input = Input
        self.wb = load_workbook(filename=Input)

        ws = self.wb.active  # should always activate first sheet
        A2 = ws["A2"].value
        if A2 not in known_modules:
            raise ValueError(f"Error: A2 is not a known module ´{A2}´")

        if cmd == "down":
            self.down()
        elif cmd == "up":
            self.up()
        else:
            raise ValueError(f"ERROR: Unknown command ´{cmd}´")
        self.wb.save(Input)  # overwriting original excel

    def down(self) -> None:
        """
        Given a properly prepared Excel sheet, du will fill in the requested
        data from RIA

        Expects
        * data in self.wb (side effect)

        Returns
        * nothing, but writes to self.wb (side effect)

        Decisions
        * Requires one or more queries to RIA? Let's try with one first.
        * At this stage we will be aggressively writing file caches for
          everything for debugging purposes

        Steps
        * we parse the first row looking for fields that need filling in
        * we create a query, save it, execute it
        * we receive the response and save it
        * we use the response to fill in the requested data
        """

        ws = self.wb.active  # should always activate first sheet
        mtype = ws["A2"].value
        now = datetime.now()
        ws["B1"] = now.isoformat()
        ws["B1"].font = Font(color="005655")

        query = Search(module=mtype)
        if ws.max_row < 3:
            raise ValueError("Error: Excel input has nothing to work on!")
        if ws.max_row > 3:
            query.OR()
        IDs = f"A3:A{ws.max_row}"
        for col in ws[IDs]:  # .iter_cols(min_row=3, max_col=1)
            for cell in col:
                print(f"{cell.row} {mtype} {cell.value}")
                query.addCriterion(
                    operator="equalsField", field="__id", value=str(cell.value)
                )

        requested = self.requestedFields()
        print(requested)
        for colNo in requested:
            query.addField(field=requested[colNo]["field"])

        query.validate(mode="search")
        query.toFile(path="query.debug.xml")
        print("Executing query...")
        m = self.sar.search(query=query)
        m.toFile(path="response.debug.xml")

        for colNo in requested:
            print(f"FF {colNo}")
            for item in m.iter(module=mtype):
                ID = int(item.xpath("@id")[0])
                field = requested[colNo]["field"]
                value = item.xpath(
                    f"m:dataField[@name = '{field}']/m:value/text()", namespaces=NSMAP
                )[0]
                self.fillIn(ID=ID, value=value, colNo=colNo)

    def fillIn(self, *, ID: int, value: str, colNo: int) -> None:
        """
        For a given ID, write value {value} to column no {colNo}
        """
        ws = self.wb.active  # should always activate first sheet
        print(f"filling in {ID} {value} into column no {colNo}")

        # we're calling "A4" a cell name

        colL = openpyxl.utils.cell.get_column_letter(colNo)
        rno = self._rowById(ID=ID)
        cname = f"{colL}{rno}"
        ws[cname] = value
        # I am trying to color code the different stages,
        # but I doubt user will be able to mark their changes as well
        ws[cname].font = Font(color="005655")

    def requestedFields(self) -> dict:
        """
        Atm, we're only working on cols from the Excel which start with
        dataField
        later we need to search for these fields in the response and
        write into exactly those columns, so we have to "remember" the fields we picked

        {
            1: {  # colNo as int, 1-based
                "field": "ObjTechnicalTermClb",  # aka zcName
                "full": "dataField.ObjTechnicalTermClb",
                "param": "value",
            }
        }
        """

        ws = self.wb.active  # should always activate first sheet
        requested = dict()
        colL = openpyxl.utils.cell.get_column_letter(ws.max_column)
        labels = f"B2:{colL}2"  # field labels
        for row in ws[labels]:
            for cell in row:
                # print (f"LLL:{labels} {cell} {cell.value}")
                if cell.value is not None and cell.value.startswith("dataField"):
                    field = cell.value.split(".", maxsplit=1)[1]
                    requested[cell.column] = {
                        "field": field,
                        "full": cell.value,
                    }
                    print(f"field {cell.column} {field}")
        if len(requested) < 1:
            raise ValueError("Error: No fields requested from Excel input file")
        return requested

    def up(self) -> None:
        """
        Loop through the first sheet at self.wb and upload changes to RIA

        We should try to avoid atomic changes where every changed field
        results in one upload. Let's try to loop thru row by row so that all
        changes in one row become one upload.
        """
        ws = self.wb.active  # should always activate first sheet
        requested = self.requestedFields()
        mtype = ws["A2"].value
        for row in ws.iter_rows(min_row=3):
            ID = row[0].value
            rno = row[0].row
            for colNo in requested:
                colL = openpyxl.utils.cell.get_column_letter(colNo)
                cname = f"{colL}{rno}"
                value = ws[cname].value
                field = requested[colNo]["field"]
                print(f"{mtype} {ID} {cname} : {ws[cname].value}")
                # This is the atomic per field approach that I wanted to avoid
                self._updateField(mtype=mtype, ID=ID, dataField=field, value=value)

    #
    # private
    #

    def _updateField(self, *, mtype: str, ID: int, dataField: str, value: str) -> None:
        """
        If I write the same Sachbegriff to a record that already exists,
        the timestamp gets updated, but the log entry has no entry. I
        usually want to avoid that.

        Also: don't change the online record (back to a old state) if it has
        been changed since the download of the data to Excel.

        Potentially, I could do some of these tests earlier, so I dont have
        to repeat them for every field in the same row.
        """
        q = Search(module=mtype)
        q.addCriterion(operator="equalsField", field="__id", value=str(ID))
        q.addField(field="__lastModified")
        q.addField(field=dataField)
        q.validate(mode="search")
        # q.toFile(path="query.debug.xml")
        m = self.sar.search(query=q)
        m.toFile(path="check.debug.xml")
        item = m[mtype, ID]  # type: ignore
        lastMod = dateutil.parser.isoparse(
            item.xpath(
                "m:systemField[@name = '__lastModified']/m:value/text()",
                namespaces=NSMAP,
            )[0]
        )
        try:
            onlineValue = item.xpath(
                f"m:dataField[@name = '{dataField}']/m:value/text()",
                namespaces=NSMAP,
            )[0]
        except:
            onlineValue = None
        ws = self.wb.active
        downloadTime = ws["B1"].value

        # print (f"oldValue {oldValue} {lastMod}")
        if onlineValue != value:
            # Excel has different value than online RIA
            print(f"INFO: online {onlineValue} <-> Excel {value}")
            if lastMod < downloadTime:
                # online RIA entry is older than Excel download
                print("INFO: Excel is newer than online data, update required")
                self.api.updateField2(
                    mtype=mtype, ID=ID, dataField=dataField, value=value
                )
            else:
                print(
                    "WARNING: Value has been changed online since download to Excel (no update)"
                )

    def _rowById(self, *, ID: int) -> int:
        ws = self.wb.active
        # I could write an index of IDs to memory to save time; but not now
        IDs = f"A3:A{ws.max_row}"
        for col in ws[IDs]:
            for cell in col:
                if cell.value == ID:
                    return cell.row
        raise ValueError("Error: ID `{ID}` not found!")
