"""
mkGrp

Take an Excel sheet as input, pick a column with IDs (objIds) and make a new 
list in RIA with those objects.

Usage
    mkGrp -i excel.xlsx -c r -a
    
"""

import argparse
from mpapi.client import MpApi
from mpapi.module import Module
from openpyxl import load_workbook

with open("credentials.py") as f:
    exec(f.read())


class GrpMaker:
    def __init__(self, *, user, pw, baseURL) -> None:
        self.client = MpApi(baseURL=baseURL, user=user, pw=pw)

    def new(self, *, column, Input, name):
        """
        <moduleReferenceItem moduleItemId="2550610">
            <dataField dataType="Long" name="SortLnu">
              <value>1</value>
            </dataField>
        </moduleReferenceItem>
        """
        wb = load_workbook(Input)
        ws = wb.active
        print(f" using sheet {ws.title} column {column}")

        # download an example
        # m = self.client.getItem2(mtype="ObjectGroup", ID=184398)
        # m.toFile(path="grp184398")

        # make a new group; very basic

        m1 = Module()
        group = m1.module(name="ObjectGroup")
        item = m1.moduleItem(parent=group)
        m1.dataField(
            parent=item, dataType="Varchar", name="OgrNameTxt", value=f"{name}"
        )
        vr = m1.vocabularyReference(
            parent=item, name="OgrTypeVoc", ID=65639, instanceName="OgrTypeVgr"
        )
        print(f"VR {vr}")
        m1.vocabularyReferenceItem(parent=vr, ID=2307936, name="groupObjects")
        # <moduleReference name="OgrObjectRef" targetModule="Object" multiplicity="M:N" size="170">
        mr = m1.moduleReference(
            parent=item, name="OgrObjectRef", targetModule="Object"
        )  # , multiplicity="M:N"

        # m1.moduleReferenceItem(parent=mr, moduleItemId=998476)
        row_count = 1
        for cell in ws[column]:
            if row_count > 1:
                print(f"object {cell.value}")
                m1.moduleReferenceItem(parent=mr, moduleItemId=int(cell.value))
            row_count += 1
        print(m1.toString())
        m2 = self.client.createItem2(mtype="ObjectGroup", data=m1)
        print(m2)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="make a group in RIA from Excel input")
    # parser.add_argument('-a', '--act', help='act or not to act', action='store_true')
    parser.add_argument(
        "-c", "--column", help="specify the excel column with IDs", required=True
    )
    parser.add_argument("-i", "--input", help="path to Excel input file", required=True)
    parser.add_argument("-n", "--name", help="name of new group", required=True)

    args = parser.parse_args()
    gm = GrpMaker(user=user, pw=pw, baseURL=baseURL)
    gm.new(
        Input=args.input, column=args.column, name=args.name
    )  # no return value planned ATM
